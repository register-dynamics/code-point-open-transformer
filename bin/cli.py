from os import listdir
from os.path import join, isfile
from sys import stdout
import csv

import click
from openpyxl import load_workbook
from tqdm import tqdm # progress bar

HEADERS_PATH = 'Doc/Code-Point_Open_Column_Headers.csv'
CODES_PATH = 'Doc/Codelist.xlsx'
DATA_PATH = 'Data/CSV'
COUNTRIES = {
  'E92000001': 'England',
  'S92000003': 'Scotland',
  'W92000004': 'Wales',
  'N92000002': 'N Ireland',
}

@click.command()
@click.argument('package_dir', type=click.Path(exists=True,
                dir_okay=True, file_okay=False))
def main(package_dir):
  """Denormalise and improve usability of Code-Point Open data"""

  headers = load_headers(join(package_dir, HEADERS_PATH))
  codes = load_codes(join(package_dir, CODES_PATH))

  new_headers = headers + ('Country', 'County', 'District', 'Ward', 'Geometry')
  writer = csv.DictWriter(stdout, fieldnames=new_headers)
  writer.writeheader()

  data_dir = join(package_dir, DATA_PATH)
  data_files = [f for f in listdir(data_dir) if isfile(join(data_dir, f))]

  for data_file in tqdm(data_files):
    with open(join(data_dir, data_file), newline='') as data_file:
      reader = csv.reader(data_file)
      for row in reader:
        data = dict(zip(headers, row))
        new_data = {
          **data,
          'Postcode': format_postcode(data['Postcode']),
          'Country': codes['countries'].get(data['Country_code']),
          'County': codes['counties'].get(data['Admin_county_code']),
          'District': codes['districts'].get(data['Admin_district_code']),
          'Ward': codes['wards'].get(data['Admin_ward_code']),
          'Geometry': format_geometry(data['Eastings'], data['Northings']),
        }
        writer.writerow(new_data)

def load_headers(headers_path):
  with open(headers_path, newline='') as headers_file:
    data = list(csv.reader(headers_file))
    return tuple(data[1])

def load_codes(codes_path):
  codes = {'counties': {}}
  workbook = load_workbook(codes_path)

  codes = {
    'countries': COUNTRIES,
    'counties': get_dict_from_sheet(workbook['CTY']),
    'districts': {
      **get_dict_from_sheet(workbook['DIS']),
      **get_dict_from_sheet(workbook['LBO']),
    },
    'wards': {
      **get_dict_from_sheet(workbook['DIW']),
      **get_dict_from_sheet(workbook['LBW']),
    },
  }

  return codes

def get_dict_from_sheet(sheet):
  data = dict()

  for (value, key) in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
    data[key] = value

  return data

def format_postcode(postcode):
  outward = postcode[:-3].strip()
  inward = postcode[-3:].strip()
  return '{} {}'.format(outward, inward)

def format_geometry(eastings, northings):
  return 'POINT ({} {})'.format(eastings, northings)

if __name__ == '__main__':
  main()
